# app/views_reportes.py

import datetime
from decimal import Decimal
from collections import defaultdict

from django.db.models import (
    Sum, F, DecimalField, ExpressionWrapper, Count, Min, Max,
    OuterRef, Subquery, Value, CharField
)
from django.db.models.functions import Concat
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from .models import (
    Compra, Detalle_Compra,
    Servicio, Categoria_Servicio, Detalle_Servicio,
    Producto, Categoria,
    Usuario, Comprobante_Pago
)

# =========================
# Utilidades / Helpers
# =========================

def _parse_dates_optional(request):
    """
    Devuelve (desde, hasta) si ambas fechas vienen bien formadas y desde<=hasta.
    Si ambas están vacías => (None, None) => exporta TODO.
    Si viene solo una o el orden es inválido => (None, None) => exporta TODO.
    """
    desde_str = request.GET.get('desde')
    hasta_str = request.GET.get('hasta')

    if not desde_str and not hasta_str:
        return None, None

    if (desde_str and not hasta_str) or (hasta_str and not desde_str):
        return None, None

    d = parse_date(desde_str) if desde_str else None
    h = parse_date(hasta_str) if hasta_str else None

    if d and h and h >= d:
        return d, h

    return None, None


def _money(expr):
    return ExpressionWrapper(expr, output_field=DecimalField(max_digits=14, decimal_places=2))


# ----- Formato monetario (preview y PDF) -----

def _fmt_money(val):
    """
    Devuelve el valor monetario como $X,XXX.XX.
    Acepta None o Decimal/float/int.
    """
    try:
        num = float(val or 0)
    except Exception:
        return str(val) if val is not None else ""
    return f"${num:,.2f}"

def _detect_money_cols(headers):
    """
    Detecta columnas monetarias por el header:
    si contiene '$' o palabras 'monto'/'total' (case-insensitive).
    """
    money_cols = set()
    for idx, h in enumerate(headers):
        hlow = str(h).lower()
        if "$" in str(h) or "monto" in hlow or "total" in hlow:
            money_cols.add(idx)
    return money_cols

def _format_money_rows(headers, rows):
    """
    Devuelve nuevas filas con columnas monetarias formateadas.
    """
    money_cols = _detect_money_cols(headers)
    out = []
    for r in rows:
        r2 = []
        for i, v in enumerate(r):
            if i in money_cols:
                r2.append(_fmt_money(v))
            else:
                r2.append(v if v is not None else "")
        out.append(tuple(r2))
    return out


# =========================
# Render / Exports
# =========================

def _render_inline_html(title, headers, rows, extra=None):
    """
    Vista previa HTML clara (card blanca, encabezado suave morado, filas alternadas)
    con tipografía moderna consistente con la UI del admin.
    Formatea columnas monetarias y resalta la fila "TOTAL GENERAL" si existe.
    """
    # Colores
    accent = "#6C2DC7"
    page_bg = "#1E2130"   # fondo suave oscuro de la vista
    card_bg = "#FFFFFF"
    head_bg = "#F4F0FE"   # morado muy claro
    row_a = "#FFFFFF"
    row_b = "#FAFAFD"
    text = "#1F2937"
    sub = "#6B7280"
    line = "#E5E7EB"

    # Formatear moneda
    rows_fmt = _format_money_rows(headers, rows)

    ths = "".join(
        f"<th style='padding:.75rem 1rem; border-bottom:1px solid {line}; text-align:left; font-weight:600'>{h}</th>"
        for h in headers
    )

    trs = []
    for i, r in enumerate(rows_fmt):
        bg_row = row_a if i % 2 == 0 else row_b
        # Total general en negrita
        is_total = (str(r[0]).strip().upper() == "TOTAL GENERAL")
        tds = "".join(
            f"<td style='padding:.65rem 1rem; border-bottom:1px solid {line}; color:{text};"
            f"{'font-weight:700' if is_total else ''}'>{'' if v is None else v}</td>"
            for v in r
        )
        # Fondo levemente distinto si es total
        row_style = f"background:{'#EEE5FF' if is_total else bg_row}"
        trs.append(f"<tr style='{row_style}'>{tds}</tr>")

    body = "".join(trs) if trs else f"<tr><td colspan='{len(headers)}' style='padding:1rem;color:{text}'>Sin datos.</td></tr>"
    extra_html = f"<div style='margin-left:10px'>{extra}</div>" if extra else ""

    html = f"""
    <div style="background:{page_bg};min-height:100vh;padding:16px 16px 60px 16px;
                font-family: Inter, Segoe UI, Roboto, system-ui, -apple-system, Arial, sans-serif;">
      <div style="max-width:1200px;margin:0 auto;background:{card_bg};border:1px solid {line};
                  border-radius:14px;box-shadow:0 10px 24px rgba(17,24,39,.10);overflow:hidden">
        <div style="padding:16px 18px;border-bottom:1px solid {line};display:flex;justify-content:space-between;align-items:center">
          <div>
            <h3 style="margin:0;color:{text}">{title}</h3>
            <div style="font-size:.9rem;color:{sub};margin-top:.25rem">Vista previa (no imprimible)</div>
          </div>
          <div>{extra_html}</div>
        </div>

        <div style="overflow:auto">
          <table style="width:100%;border-collapse:collapse;">
            <thead style="background:{head_bg};">
              <tr>{ths}</tr>
            </thead>
            <tbody>{body}</tbody>
          </table>
        </div>
      </div>
    </div>
    """
    return HttpResponse(html)


def _export_xlsx(filename, headers, rows):
    """Genera un XLSX con openpyxl (números crudos, ideal para análisis en Excel)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return HttpResponse("openpyxl no está instalado.", status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # headers
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)

    # rows (sin formateo de moneda para poder sumar en Excel)
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # auto width
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


# ===== Encabezado PDF tipo comprobante (ReportLab) =====
def _export_pdf_reportlab(filename, title, headers, rows, subtitle=None):
    """
    Genera un PDF con encabezado profesional (logo, marca, título, rango y paginación)
    usando ReportLab. Si ReportLab no está instalado, retorna None.
    Aplica formato monetario y resalta la fila 'TOTAL GENERAL' si existe.
    """
    try:
        import io, os
        from django.conf import settings
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.pdfgen.canvas import Canvas
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from django.utils import timezone
    except Exception:
        return None  # No hay reportlab

    # Paleta
    MORADO = colors.HexColor("#6C2DC7")
    TEXTO  = colors.HexColor("#212529")
    SEC    = colors.HexColor("#6C757D")
    LINE   = colors.HexColor("#E9ECEF")
    TOTAL_BG = colors.HexColor("#EEE5FF")

    # Formatear moneda en data
    rows_fmt = _format_money_rows(headers, rows)

    buf = io.BytesIO()
    width, height = A4
    left_margin = right_margin = 18 * mm
    top_margin = 18 * mm
    bottom_margin = 16 * mm

    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=left_margin,
        rightMargin=right_margin,
        topMargin=top_margin + 52,
        bottomMargin=bottom_margin + 20
    )

    def draw_header_footer(canvas: Canvas, _doc):
        canvas.setFillColor(colors.white)
        canvas.rect(0, 0, width, height, fill=1, stroke=0)

        # Logo (opcional)
        try:
            logo_path = os.path.join(settings.BASE_DIR, 'static', 'img', 'logo.png')
            if os.path.exists(logo_path):
                logo = ImageReader(logo_path)
                canvas.drawImage(logo, left_margin, height - 40*mm, width=26*mm, height=26*mm, mask='auto')
        except Exception:
            pass

        # Marca + Título
        canvas.setFont("Helvetica-Bold", 15)
        canvas.setFillColor(MORADO)
        canvas.drawString(left_margin + 30*mm, height - 18*mm, "FLORISTERÍA FLOWERS TODAY")

        canvas.setFont("Helvetica-Bold", 12)
        canvas.setFillColor(TEXTO)
        canvas.drawString(left_margin + 30*mm, height - 26*mm, str(title))

        # Subtítulo (rango) + fecha de generación
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(SEC)
        if subtitle:
            canvas.drawString(left_margin + 30*mm, height - 32*mm, subtitle)
        gen = timezone.localtime(timezone.now()).strftime("Generado: %d/%m/%Y %H:%M")
        canvas.drawRightString(width - right_margin, height - 18*mm, gen)

        # Línea separadora
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(1)
        canvas.line(left_margin, height - 36*mm, width - right_margin, height - 36*mm)

        # Pie: número de página
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(SEC)
        canvas.drawRightString(width - right_margin, bottom_margin - 6, f"Página {canvas.getPageNumber()}")

    # Data
    data = [headers] + rows_fmt

    # Cálculo simple de anchos de columna
    def _col_widths(tbl_data, total_width):
        lens = [max(len(str(r[i])) if i < len(r) and r[i] is not None else 0 for r in tbl_data)
                for i in range(len(tbl_data[0]))]
        total = sum(lens) or 1
        widths = []
        for l in lens:
            pct = max(0.12, (l / total))
            widths.append(pct)
        s = sum(widths)
        widths = [w / s for w in widths]
        return [total_width * w for w in widths]

    table_width = width - (left_margin + right_margin)
    col_widths = _col_widths(data, table_width)

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # Estilo base
    style = [
        # Head
        ('BACKGROUND', (0, 0), (-1, 0), MORADO),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10.5),
        ('TOPPADDING', (0, 0), (-1, 0), 7),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
        ('LINEBELOW', (0, 0), (-1, 0), 1, MORADO),

        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9.5),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor("#FAFAFD")]),  # hasta penúltima
        ('GRID', (0, 0), (-1, -1), 0.25, LINE),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
    ]

    # Buscar la fila "TOTAL GENERAL" y destacarla
    total_row_index = None
    for idx in range(1, len(data)):
        if str(data[idx][0]).strip().upper() == "TOTAL GENERAL":
            total_row_index = idx
            break

    if total_row_index is not None:
        style += [
            ('BACKGROUND', (0, total_row_index), (-1, total_row_index), TOTAL_BG),
            ('FONTNAME', (0, total_row_index), (-1, total_row_index), 'Helvetica-Bold'),
            ('LINEABOVE', (0, total_row_index), (-1, total_row_index), 0.75, MORADO),
        ]

    table.setStyle(TableStyle(style))

    doc.build([table], onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)

    pdf_data = buf.getvalue()
    buf.close()
    resp = HttpResponse(pdf_data, content_type="application/pdf")
    resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    return resp


def _export_pdf(filename, title, headers, rows):
    """
    Fallback PDF: intenta WeasyPrint y luego xhtml2pdf.
    Aplica formato monetario y, si ninguno está, vuelve a la preview HTML con CTA Excel.
    """
    # Formatear moneda para HTML que alimenta a los motores
    rows_fmt = _format_money_rows(headers, rows)

    accent = "#6C2DC7"
    table_head = "".join(
        f"<th style='padding:6px 10px;text-align:left;border-bottom:1px solid #ccc'>{h}</th>"
        for h in headers
    )
    table_rows = "".join(
        "<tr>" + "".join(
            f"<td style='padding:6px 10px;border-bottom:1px solid #eee'>{'' if v is None else v}</td>"
            for v in r
        ) + "</tr>"
        for r in rows_fmt
    )
    html_doc = f"""
    <html><head><meta charset="utf-8"><title>{title}</title></head>
    <body style="font-family: Inter, Segoe UI, Roboto, Arial, sans-serif; font-size:12px">
      <h2 style="color:{accent}; margin:0 0 8px 0;">{title}</h2>
      <table style="width:100%;border-collapse:collapse">
        <thead style="background:{accent}20"><tr>{table_head}</tr></thead>
        <tbody>{table_rows}</tbody>
      </table>
    </body></html>
    """

    # 1) WeasyPrint
    try:
        from weasyprint import HTML, CSS
        pdf = HTML(string=html_doc).write_pdf(stylesheets=[CSS(string="body{font-size:12px}")])
        resp = HttpResponse(pdf, content_type="application/pdf")
        resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return resp
    except Exception:
        pass

    # 2) xhtml2pdf (pisa)
    try:
        from xhtml2pdf import pisa
        from io import BytesIO
        pdf_io = BytesIO()
        pisa.CreatePDF(html_doc, dest=pdf_io)
        pdf_io.seek(0)
        resp = HttpResponse(pdf_io.getvalue(), content_type="application/pdf")
        resp['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return resp
    except Exception:
        # 3) Fallback: preview con CTA Excel
        cta = f"""
        <a href="?format=xlsx" style="background:{accent};color:white;text-decoration:none;padding:.55rem .9rem;border-radius:10px;box-shadow:0 6px 16px rgba(108,45,199,.3);">
          Descargar Excel
        </a>
        """
        return _render_inline_html(f"[PREVIEW PDF] {title}", headers, rows_fmt, extra=cta)


# =========================
# Consultas / Reportes
# =========================

def q_ventas_por_producto(desde, hasta):
    """Unidades y monto por producto (solo comprobantes Pa)."""
    importe = _money(F('precio_unitario_compra') * F('cantidad_producto_compra'))
    qs = Detalle_Compra.objects.filter(id_compra__comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        qs = qs.filter(id_compra__fecha_compra__range=(desde, hasta))
    qs = (qs.values('id_producto__nombre_producto', 'id_producto__id_categoria__nombre_categoria')
            .annotate(unidades=Sum('cantidad_producto_compra'), monto=Sum(importe))
            .order_by('-monto', '-unidades', 'id_producto__nombre_producto'))

    headers = ["Producto", "Categoría", "Unidades", "Monto ($)"]
    rows = [
        (r['id_producto__nombre_producto'],
         r['id_producto__id_categoria__nombre_categoria'],
         r['unidades'] or 0,
         float(r['monto'] or 0))
        for r in qs
    ]
    # TOTAL GENERAL
    tot_u = sum(r[2] or 0 for r in rows)
    tot_m = sum(r[3] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", "", tot_u, tot_m))
    return headers, rows


def q_ventas_por_categoria(desde, hasta):
    """Unidades y monto agrupado por categoría (solo comprobantes Pa)."""
    importe = _money(F('precio_unitario_compra') * F('cantidad_producto_compra'))
    qs = Detalle_Compra.objects.filter(id_compra__comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        qs = qs.filter(id_compra__fecha_compra__range=(desde, hasta))
    qs = (qs.values('id_producto__id_categoria__nombre_categoria')
            .annotate(unidades=Sum('cantidad_producto_compra'), monto=Sum(importe))
            .order_by('-monto'))

    headers = ["Categoría", "Unidades", "Monto ($)"]
    rows = [
        (r['id_producto__id_categoria__nombre_categoria'],
         r['unidades'] or 0,
         float(r['monto'] or 0))
        for r in qs
    ]
    # TOTAL GENERAL
    tot_u = sum(r[1] or 0 for r in rows)
    tot_m = sum(r[2] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", tot_u, tot_m))
    return headers, rows



def q_ventas_por_servicio_detalle(desde, hasta):
    """Listado de servicios vendidos (monto y cantidad) solo servicios 'Ac'."""
    importe = _money(F('precio_unitario_servicio') * F('cantidad_producto_servicio'))
    qs = Detalle_Servicio.objects.filter(id_servicio__estado_servicio='Ac')
    if desde and hasta:
        qs = qs.filter(id_servicio__fecha_servicio__range=(desde, hasta))
    qs = (qs.values('id_servicio',
                    'id_servicio__descripcion_servicio',
                    'id_servicio__id_categoria_servicio__nombre_categoria_servicio',
                    'id_servicio__fecha_servicio')
            .annotate(items=Sum('cantidad_producto_servicio'), total=Sum(importe))
            .order_by('-total', '-items', 'id_servicio'))

    headers = ["Servicio", "Categoría Servicio", "Fecha", "Ítems", "Total ($)"]
    rows = [
        (r['id_servicio__descripcion_servicio'],
         r['id_servicio__id_categoria_servicio__nombre_categoria_servicio'],
         r['id_servicio__fecha_servicio'].isoformat() if r['id_servicio__fecha_servicio'] else "",
         r['items'] or 0,
         float(r['total'] or 0))
        for r in qs
    ]
    # TOTAL GENERAL
    tot_items = sum(r[3] or 0 for r in rows)
    tot_monto = sum(r[4] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", "", "", tot_items, tot_monto))
    return headers, rows

def q_ventas_totales_servicios(desde, hasta):
    """Totales por categoría de servicio (solo 'Ac')."""
    importe = _money(F('precio_unitario_servicio') * F('cantidad_producto_servicio'))
    qs = Detalle_Servicio.objects.filter(id_servicio__estado_servicio='Ac')
    if desde and hasta:
        qs = qs.filter(id_servicio__fecha_servicio__range=(desde, hasta))
    qs = (qs.values('id_servicio__id_categoria_servicio__nombre_categoria_servicio')
            .annotate(servicios=Count('id_servicio', distinct=True),
                      items=Sum('cantidad_producto_servicio'),
                      total=Sum(importe))
            .order_by('-total'))

    headers = ["Categoría Servicio", "Servicios Distintos", "Unidades/Ítems", "Total ($)"]
    rows = [
        (r['id_servicio__id_categoria_servicio__nombre_categoria_servicio'],
         r['servicios'] or 0,
         r['items'] or 0,
         float(r['total'] or 0))
        for r in qs
    ]
    # TOTAL GENERAL
    tot_srv = sum(r[1] or 0 for r in rows)
    tot_items = sum(r[2] or 0 for r in rows)
    tot_monto = sum(r[3] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", tot_srv, tot_items, tot_monto))
    return headers, rows



def q_ventas_totales_productos(desde, hasta):
    """Un único total del rango (solo comprobantes Pa)."""
    importe = _money(F('precio_unitario_compra') * F('cantidad_producto_compra'))
    qs = Detalle_Compra.objects.filter(id_compra__comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        qs = qs.filter(id_compra__fecha_compra__range=(desde, hasta))
    agg = qs.aggregate(unidades=Sum('cantidad_producto_compra'), monto=Sum(importe))
    rango = (f"{desde.isoformat()} → {hasta.isoformat()}" if (desde and hasta) else "Todo")
    headers = ["Rango", "Unidades Totales", "Monto Total ($)"]
    rows = [(rango, agg['unidades'] or 0, float(agg['monto'] or 0))]
    return headers, rows


def q_cantidad_productos_vendidos(desde, hasta):
    """Conteo por producto (solo comprobantes Pa)."""
    qs = Detalle_Compra.objects.filter(id_compra__comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        qs = qs.filter(id_compra__fecha_compra__range=(desde, hasta))
    qs = (qs.values('id_producto__nombre_producto')
            .annotate(unidades=Sum('cantidad_producto_compra'))
            .order_by('-unidades', 'id_producto__nombre_producto'))
    headers = ["Producto", "Unidades"]
    rows = [(r['id_producto__nombre_producto'], r['unidades'] or 0) for r in qs]
    return headers, rows


def q_existencias_estado():
    """Stock actual + mínimos/máximos + activo."""
    qs = (Producto.objects
          .values('nombre_producto', 'existencia_producto', 'cantidad_minima', 'cantidad_maxima',
                  'producto_activo', 'id_categoria__nombre_categoria')
          .order_by('nombre_producto'))
    headers = ["Producto", "Categoría", "Existencia", "Mín.", "Máx.", "Activo"]
    rows = [
        (r['nombre_producto'], r['id_categoria__nombre_categoria'], r['existencia_producto'],
         r['cantidad_minima'], r['cantidad_maxima'],
         "Sí" if r['producto_activo'] else "No")
        for r in qs
    ]
    return headers, rows


def q_clientes():
    """Todos los usuarios activos (si luego defines rol cliente, se filtra)."""
    qs = (Usuario.objects
          .filter(usuario_activo=True)
          .values('nombre_usuario', 'apellido_usuario', 'correo_usuario',
                  'telefono_usuario', 'id_rol__nombre_rol')
          .order_by('apellido_usuario', 'nombre_usuario'))
    headers = ["Nombre", "Apellido", "Correo", "Teléfono", "Rol"]
    rows = [
        (r['nombre_usuario'], r['apellido_usuario'], r['correo_usuario'],
         r['telefono_usuario'], r['id_rol__nombre_rol'])
        for r in qs
    ]
    return headers, rows


def q_comparativo_mensual(desde, hasta):
    """
    Suma por mes (YYYY-MM) SOLO Pa. Si no hay fechas, toma todo el rango disponible Pa.
    """
    base = Compra.objects.filter(comprobante_pago__estado_comprobante='Pa')

    if desde and hasta:
        start = datetime.date(desde.year, desde.month, 1)
        end_first = datetime.date(hasta.year, hasta.month, 1)
        qs = base.filter(fecha_compra__range=(start, hasta))
    else:
        agg = base.aggregate(minf=Min('fecha_compra'), maxf=Max('fecha_compra'))
        if not agg['minf'] or not agg['maxf']:
            headers = ["Mes (YYYY-MM)", "Total ($)"]
            return headers, []
        start = datetime.date(agg['minf'].year, agg['minf'].month, 1)
        end_first = datetime.date(agg['maxf'].year, agg['maxf'].month, 1)
        qs = base

    by_month = defaultdict(Decimal)
    for c in qs:
        key = f"{c.fecha_compra.year}-{c.fecha_compra.month:02d}"
        by_month[key] += Decimal(c.total_compra or 0)

    months = []
    cursor = start
    while cursor <= end_first:
        months.append(f"{cursor.year}-{cursor.month:02d}")
        cursor = (datetime.date(cursor.year + 1, 1, 1)
                  if cursor.month == 12 else
                  datetime.date(cursor.year, cursor.month + 1, 1))

    headers = ["Mes (YYYY-MM)", "Total ($)"]
    rows = [(m, float(by_month.get(m, 0))) for m in months]

    # TOTAL GENERAL
    tot = sum(r[1] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", tot))
    return headers, rows


def q_ventas_detalle(desde, hasta):
    """
    Ventas (detalle por compra) SOLO Pa:
    Fecha, Comprobante, Cliente, Monto (+ Total general).
    """
    compras = Compra.objects.filter(comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        compras = compras.filter(fecha_compra__range=(desde, hasta))

    # último comprobante por compra
    ult_comp = (Comprobante_Pago.objects
                .filter(id_compra=OuterRef('pk'))
                .order_by('-fecha_comprobante')
                .values('codigo_comprobante')[:1])

    qs = (compras
          .annotate(
              comprobante=Subquery(ult_comp, output_field=CharField()),
              cliente=Concat(
                  F('id_usuario__nombre_usuario'),
                  Value(' '),
                  F('id_usuario__apellido_usuario'),
                  output_field=CharField()
              )
          )
          .values('fecha_compra', 'comprobante', 'cliente', 'total_compra')
          .order_by('-fecha_compra', '-id_compra'))

    headers = ["Fecha", "Comprobante", "Cliente", "Monto ($)"]
    rows = [
        (
            r['fecha_compra'].isoformat() if r['fecha_compra'] else "",
            r['comprobante'] or "-",
            r['cliente'] or "",
            float(r['total_compra'] or 0)
        )
        for r in qs
    ]
    # TOTAL GENERAL
    tot = sum(r[3] or 0.0 for r in rows)
    if rows:
        rows.append(("TOTAL GENERAL", "", "", tot))
    return headers, rows


def q_ventas_totales(desde, hasta):
    """
    Totales del rango o global:
    - Productos: Detalle_Compra solo Pa (fecha de compra)
    - Servicios: Detalle_Servicio solo 'Ac' (fecha del servicio)
    + Fila final "TOTAL GENERAL".
    """
    # Productos (Pa)
    importe_prod = _money(F('precio_unitario_compra') * F('cantidad_producto_compra'))
    dc = Detalle_Compra.objects.filter(id_compra__comprobante_pago__estado_comprobante='Pa')
    if desde and hasta:
        dc = dc.filter(id_compra__fecha_compra__range=(desde, hasta))
    agg_p = dc.aggregate(unidades=Sum('cantidad_producto_compra'),
                         monto=Sum(importe_prod))

    # Servicios (Ac)
    importe_srv = _money(F('precio_unitario_servicio') * F('cantidad_producto_servicio'))
    ds = Detalle_Servicio.objects.filter(id_servicio__estado_servicio='Ac')
    if desde and hasta:
        ds = ds.filter(id_servicio__fecha_servicio__range=(desde, hasta))
    agg_s = ds.aggregate(unidades=Sum('cantidad_producto_servicio'),
                         monto=Sum(importe_srv))

    u_prod = agg_p["unidades"] or 0
    m_prod = float(agg_p["monto"] or 0)
    u_serv = agg_s["unidades"] or 0
    m_serv = float(agg_s["monto"] or 0)

    headers = ["Tipo", "Unidades/Ítems", "Monto ($)"]
    rows = [
        ("Productos", u_prod, m_prod),
        ("Servicios", u_serv, m_serv),
        ("TOTAL GENERAL", u_prod + u_serv, m_prod + m_serv),
    ]
    return headers, rows



# =========================
# Controlador principal
# =========================

def exportar_reporte(request, endpoint: str):
    fmt = (request.GET.get('format') or 'html').lower()

    # Endpoints SIN fechas
    if endpoint in ('existencias_estado', 'clientes'):
        headers, rows = (q_existencias_estado() if endpoint == 'existencias_estado' else q_clientes())
        title = "Existencias y estado de productos" if endpoint == 'existencias_estado' else "Listado de clientes"
        filename = endpoint

        if fmt == 'xlsx':
            return _export_xlsx(filename, headers, rows)
        elif fmt == 'pdf':
            # Encabezado pro (ReportLab). Si no está, fallback a _export_pdf()
            resp = _export_pdf_reportlab(filename, title, headers, rows, subtitle=None)
            if resp is not None:
                return resp
            return _export_pdf(filename, title, headers, rows)
        else:
            return _render_inline_html(title, headers, rows)

    # Endpoints CON fechas (opcionales)
    desde, hasta = _parse_dates_optional(request)

    if endpoint == 'ventas_por_producto':
        headers, rows = q_ventas_por_producto(desde, hasta)
        title = f"Ventas por producto ({desde} a {hasta})" if (desde and hasta) else "Ventas por producto (Todo)"
    elif endpoint == 'ventas_por_categoria':
        headers, rows = q_ventas_por_categoria(desde, hasta)
        title = f"Ventas por categoría ({desde} a {hasta})" if (desde and hasta) else "Ventas por categoría (Todo)"
    elif endpoint == 'ventas_por_servicio_detalle':
        headers, rows = q_ventas_por_servicio_detalle(desde, hasta)
        title = f"Servicios vendidos (detalle) ({desde} a {hasta})" if (desde and hasta) else "Servicios vendidos (detalle) (Todo)"
    elif endpoint == 'ventas_totales_servicios':
        # compatibilidad si aún lo llamas, pero mejor usar 'ventas_totales'
        headers, rows = q_ventas_totales_servicios(desde, hasta)
        title = f"Ventas totales por servicios ({desde} a {hasta})" if (desde and hasta) else "Ventas totales por servicios (Todo)"
    elif endpoint == 'ventas_totales_productos':
        # compatibilidad si aún lo llamas, pero mejor usar 'ventas_totales'
        headers, rows = q_ventas_totales_productos(desde, hasta)
        title = f"Ventas totales de productos ({desde} a {hasta})" if (desde and hasta) else "Ventas totales de productos (Todo)"
    elif endpoint == 'ventas_totales':
        headers, rows = q_ventas_totales(desde, hasta)
        title = (f"Ventas totales (Productos y Servicios) ({desde} a {hasta})"
                 if (desde and hasta) else "Ventas totales (Productos y Servicios) (Todo)")
    elif endpoint == 'cantidad_productos_vendidos':
        headers, rows = q_cantidad_productos_vendidos(desde, hasta)
        title = f"Cantidad de productos vendidos ({desde} a {hasta})" if (desde and hasta) else "Cantidad de productos vendidos (Todo)"
    elif endpoint == 'comparativo_mensual':
        headers, rows = q_comparativo_mensual(desde, hasta)
        title = f"Comparativo ventas mensuales ({desde} a {hasta})" if (desde and hasta) else "Comparativo ventas mensuales (Todo)"
    elif endpoint == 'ventas_detalle':
        headers, rows = q_ventas_detalle(desde, hasta)
        title = f"Ventas (detalle por compra) ({desde} a {hasta})" if (desde and hasta) else "Ventas (detalle por compra) (Todo)"
    else:
        return HttpResponse("Endpoint no soportado.", status=404)

    filename = endpoint
    subtitle = f"Rango: {desde} a {hasta}" if (desde and hasta) else None

    if fmt == 'xlsx':
        return _export_xlsx(filename, headers, rows)
    elif fmt == 'pdf':
        # 1) Encabezado tipo comprobante con ReportLab
        resp = _export_pdf_reportlab(filename, title, headers, rows, subtitle=subtitle)
        if resp is not None:
            return resp
        # 2) Fallback a motores existentes (WeasyPrint/xhtml2pdf) o preview
        return _export_pdf(filename, title, headers, rows)
    else:
        return _render_inline_html(title, headers, rows)
